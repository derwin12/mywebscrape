from dataclasses import dataclass
from openpyxl import load_workbook
import os
import urllib.parse

from my_funcs import insert_sequence

storename = 'UXSG'
BASEURL = "https://www.facebook.com/groups/228562318717861/search/?q="

filename = 'Unofficial xLights Sharing Group Shared Files List (31 Jan 2022).xlsx'
name_column = 6
artist_column = 7
author_column = 9


@dataclass
class Sequence:
    name: str
    url: str


def main() -> None:
    print(f"Loading UXSG Spreadsheet")
    dir = os.getcwd() + "\\..\\app\\Data\\"
    wb = load_workbook(filename = dir + filename, data_only=True)
    sheet_obj = wb.active
    sequences = []

    for row in range(5,1000):
        if sheet_obj.cell(column=2, row=row).value and \
            "_NA" not in sheet_obj.cell(column=name_column, row=row).value:
                # exlude the NA ones
                sequence_name = sheet_obj.cell(column=name_column, row=row).value + " " + \
                    "-- " + sheet_obj.cell(column=artist_column, row=row)._value + \
                    " (" + sheet_obj.cell(column=author_column, row=row).value + ")"
                product_url = BASEURL + urllib.parse.quote(sheet_obj.cell(column=name_column, row=row).value.strip() +\
                    " " + sheet_obj.cell(column=artist_column, row=row)._value)

                sequences.append(Sequence(sequence_name, product_url))

    for product in sequences:
        insert_sequence(store=storename, url=product.url, name=product.name)


if __name__ == "__main__":
    main()
